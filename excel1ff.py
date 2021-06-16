#Excel,Excel,Excel
#Creating new Excel files

from openpyxl import Workbook
import random as rnd
"""
book = Workbook()
sheet1 = book.active

sheet1.cell(row = 1, column = 1).value = "Hello"
sheet1.cell(row = 1, column = 2).value = "How are you"

book.save(filename = "file.xlsx")

wkb = Workbook()
number = wkb.active

for i in range(1,21):
    for j in range(1,7):
        number.cell(row = i, column = j).value\
                        = rnd.randint(1,20)

wkb.save(filename = "num2.xlsx")
"""
#Loading Existing excel files

from openpyxl import load_workbook

book = load_workbook("num2.xlsx")
sheet1 = book["Sheet"]

for i in range(1,7):
    print(sheet1.cell(row = i, column = 2).value)


    
#sheet1["D7"] = "Omolola"

#book.save(filename ="num.xlsx")

#print(sheet1.cell(row = 7, column = 4).value)

