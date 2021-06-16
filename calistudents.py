from faker import Faker
from openpyxl import Workbook
import random as kp
import os
global profile
global file
global sheet
profile= Faker()
file= Workbook()
sheet = file.active
def name():
    for x in range (1,10001):
        sheet.cell(row= x, column= 1 ).value=(f"NAME:\t{profile.name()}")
    file.save("calistudents.xlsx")
    
def age():
    for x in range (1,10001):
        sheet.cell(row= x, column= 2 ).value=(f"AGE:\t{kp.randint(16,30)}")
    file.save("calistudents.xlsx")
def matric_number():
    for x in range (1,1000):
        sheet.cell(row= x, column= 3 ).value=(f"MATRIC NO:\t{kp.randint(100000000,1000000000)}")
    file.save("calistudents.xlsx")
def faculty():
    fac=["Faculty of Technology","Faculty of Science","Faculty of Veterinary Medicine",
    "Faculty of Clinical Science and Dentistry","Faculty of Basic Medical Sciences",
    "Faculty of Education","Faculty of Agriculture & Forestry",
    "Faculty of Pharmacy","Faculty of Social Science",
    "Faculty of Arts",]
    for x in range (1,10001):
        #global rand
        rand=kp.choice(fac)
        sheet.cell(row= x, column= 4 ).value=(f"FACULTY:\t{rand}")
    file.save("calistudents.xlsx")
def duration():
    for x in range (1,10001):
        sheet.cell(row= x, column= 5 ).value=(f"PROGRAM DURATION:\t{kp.randint(4,6)}years")
    file.save("calistudents.xlsx")
def level():
    for x in range (1,10001):
        sheet.cell(row= x, column= 6 ).value=(f"LEVEL:\t{kp.randrange(100,501,100)}level")
    file.save("calistudents.xlsx")
def phone():
    for x in range (1,10001):
        sheet.cell(row= x, column= 7 ).value=(f"PHONE NO:\t{profile.phone_number()}")
    file.save("calistudents.xlsx")
def mail():
    for x in range (1,10001):
        sheet.cell(row= x, column= 8 ).value=(f"EMAIL:\t{profile.email()}")
    file.save("calistudents.xlsx")
def mac():
    for x in range (1,10001):
        sheet.cell(row= x, column= 9 ).value=(f"MAC ADDRESS:\t{profile.mac_address()}")
    file.save("calistudents.xlsx")
def address():
    for x in range (1,10001):
        sheet.cell(row= x, column= 10 ).value=(f"ADDRESS:\t{profile.address()}")
    file.save("calistudents.xlsx")
def Sname():
    for x in range (1,10001):
        sheet.cell(row= x, column= 11 ).value=(f"SPONSOR NAME:\t{profile.name_male()}")
    file.save("calistudents.xlsx")
def Sphone():
    for x in range (1,10001):
        sheet.cell(row= x, column= 12 ).value=(f"PHONE NUMBER:\t{profile.phone_number()}")
    file.save("calistudents.xlsx")
def job():
    for x in range (1,10001):
        job=[profile.job(),"NONE","NONE","NONE","NONE"]
        sheet.cell(row= x, column= 13 ).value=(f"JOB:\t{kp.choice(job)}")
    file.save("calistudents.xlsx")
def fee():
    for x in range (1,10001):
        sheet.cell(row= x, column= 14 ).value=(f"SEMESTER FEE:{kp.randint(100,200)}$")
    file.save("calistudents.xlsx")
def all():
    name()
    age()
    matric_number()
    faculty()
    duration()
    level()
    phone()
    mail()
    mac()
    address()
    Sname()
    Sphone()
    job()
    fee()

all()






'''def department():
    global rand
    J="Department of General Tech"

    I=["Department of Botany","Department of Microbiology",
    "Department of Industrial Chemistry","Department of Comuputer Science",
    "Department of Geology","Department of Mathematic","Department of Physics",
    "Department of Zoology"]

    A="Department of Veterinary Medicine"

    B=["Department of Physiotherapy","Department of Nursing"]

    C="Department of Biochemistry"

    D=["Department of Library,Archival Information","Department of Studies"]

    E=["Department of Agriculture","Department of Forestry","Wildlife & Fisheries"]

    F="Department of Pharmacy"

    G="Department of Geography"

    H=["Department of languages","Department of music","Department of Visual Art","Department of Talent Discovery"]
    for x in range (1,11):
        if rand == "Faculty of Technology":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{J}")
        file.save("calistudents.xlsx")
        
        if rand == "Faculty of Science":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{kp.choice(I)}")
        file.save("calistudents.xlsx")
        
        if rand == "Faculty of Veterinary Medicine":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{A}")
        file.save("calistudents.xlsx")
        
        if rand == "Faculty of Clinical Science and Dentistry":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{kp.choice(B)}")
        file.save("calistudents.xlsx")
        
        if rand == "Faculty of Basic Medical Sciences":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{C}")
        file.save("calistudents.xlsx")
        if rand == "Faculty of Education":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{kp.choice(D)}")
        file.save("calistudents.xlsx")
        if rand == "Faculty of Agriculture & Forestry":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{kp.choice(E)}")
        file.save("calistudents.xlsx")
        if rand == "Faculty of Pharmacy":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{F}")
        file.save("calistudents.xlsx")
        if rand == "Faculty of Social Science":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{G}")
        file.save("calistudents.xlsx")
        if rand == "Faculty of Arts":
            sheet.cell(row= 5, column= x ).value=(f"DEPARTMENT:\t{kp.choice(H)}")
        file.save("calistudents.xlsx")'''



        






        
    
