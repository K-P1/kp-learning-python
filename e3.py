from openpyxl import load_workbook,Workbook

from openpyxl.utils import FORMULAE
"""
Creating new excel file
wk = Workbook()
sheet1 = wk.active

sheet1.cell(row = 1, column = 1).value = 16535677
sheet1.cell(row = 1, column = 2).value = 12

wk.save(filename = "243.xlsx")

#Reading from existing file
#from openpyxl import load_workbook

book= load_workbook("mmm.xlsx")
sheet = book.active
#print(sheet.cell(row = 1, column = 1).value)
print(book.sheetnames)
#Row slicing
#print(sheet["A1:F24"])

To access elements by column, use Alphabets. Numbers to access by row

#Row and column iteration using .iter_rows,.iter_cols

for row in sheet.iter_rows():
    print(row)
"""
#for column in sheet.iter_rows():
#    print(column)


"""
Row and column iteration methods can take in the following keyword arguements
1. min_row
2. min_col
3. max_row
4. max_col
5. values_only


for column in sheet.rows:
    print(column)


#To append data to a row or column, open it, write the data you want into any cell then
    #save as new filename


To add and delete rows and columns
Use insert_row,delete_row, insert_col,delete_col
They all take max of 2 arguements
1. idx, start point of action
2. amount, row/column span
To navigate sheets, use workbook.sheetname to return a list of sheetnames
so assign workbook[f"{name of sheet}"] to get a sheet of the whole workbook

To add and remove sheet, create_sheet(name of sheet,index), remove(name of sheet)
To duplicate worksheet, copy_worksheet(name of sheet)
"""

#, to add formulas, import FORMULAE from openpyxl.utils

print(FORMULAE)



