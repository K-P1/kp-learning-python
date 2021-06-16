from openpyxl import load_workbook,Workbook

wk = Workbook()
sheet1 = wk.active

sheet1.cell(row = 1, column = 1).value = 1
sheet1.cell(row = 1, column = 2).value = 12

wk.save(filename = "24ggg.xlsx")

